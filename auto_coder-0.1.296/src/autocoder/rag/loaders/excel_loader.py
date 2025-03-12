from typing import List, Tuple
from openpyxl import load_workbook
from jinja2 import Template

def extract_text_from_excel(excel_path) -> List[Tuple[str, str]]:
    sheet_list = []
    wb = load_workbook(excel_path)
    tmpl = Template(
        """{% for row in rows %}
{% for cell in row %}"{{ cell }}"{% if not loop.last %},{% endif %}{% endfor %}{% if not loop.last %}{% endif %}{% endfor %}
        """
    )
    for ws in wb:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # 过滤掉rows中全是null的行
        rows = [row for row in rows if any(row)]
        # 所有的None都转换成空字符串
        rows = [[cell if cell is not None else "" for cell in row] for row in rows]
        content = tmpl.render(rows=rows)
        sheet_list.append([excel_path + f"#{ws.title}", content])
    wb.close()    
    return sheet_list